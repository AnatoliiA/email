from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import openpyxl


def getlastrow(tempsheet):
    """ Проверяет заполненость рядов"""
    last_row = None
    rows = list(tempsheet.iter_rows(min_col=1, max_col=1, min_row=0))
    for row in reversed(rows):
        if row[0].value is not None:
            last_row = row[0].row
            break
    return last_row if last_row else 0


def insertpass(name_company: str = "не визначено",
               phone_number_company: str = "не визначено",
               email_company: str = "не визначено",
               website: str = "не визначено",
               name_person: str = "не визначено",
               surname_person: str = "не визначено",
               phone_person: str = "не визначено",
               email_person: str = "не визначено"):
    """
    Записывает в книгу результаты отпрвки и фиксирует факт отправки

    Эта функция принимает два целочисленных аргумента и выполняет
    сложные вычисления, возвращая результат.

    Args:
       1 name_company         (str): Первый аргумент.
       2 phone_number_company (str): Второй аргумент.
       3 email_company        (str): email company
       4 website              (str): Вебсайт компании
       5 name_persons         (str): Имя персоны
       6 surname_person       (str): Фимилия персоны
       7 phone_person         (str): Телефон персоны
       8 email_person         (str): Емайл персоны
    Returns:
        возвращет True
        приошибке возвращает False
    """
    try:
        bookpath = r"utils\PythonDemo.xlsx"
        book = openpyxl.load_workbook(bookpath)
        sheet = book["sentmails"]
        current_time = datetime.now()
        # Convert the datetime object to a string
        current_time_str = current_time.strftime("%Y-%m-%d")
        try:
            line = int(getlastrow(sheet)) + 1
            sheet.cell(row=line, column=1).value = line
            sheet.cell(row=line, column=2).value = current_time_str
            sheet.cell(row=line, column=3).value = name_company
            sheet.cell(row=line, column=4).value = phone_number_company
            sheet.cell(row=line, column=5).value = email_company
            sheet.cell(row=line, column=6).value = website
            sheet.cell(row=line, column=7).value = name_person
            sheet.cell(row=line, column=8).value = surname_person
            sheet.cell(row=line, column=9).value = email_person
            sheet.cell(row=line, column=10).value = phone_person
            sheet.cell(row=line, column=11).value = "+"
        except Exception as ex:
            ex.with_traceback
            print(f"что пошло не верно {ex}")
            book.close()
            return False
        book.save("utils\PythonDemo.xlsx")
        # print(last_row)
        book.close()
        return True
    except Exception as exp:
        print(f"что пошло не верно {exp}")
        return False
